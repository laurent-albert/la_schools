# -*- coding: utf-8 -*-
# Copyright (c) 2015, LA and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import os
import babel.dates

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils.dateutils import parse_date
from frappe.utils import cint, cstr, flt, getdate, get_datetime, formatdate
from frappe.utils import get_site_name, get_site_path, get_site_base_path, get_path

import openpyxl
from openpyxl import load_workbook

class APBImportFile(Document):
	def import_abp_file(self):
		'''
			Trigger on import button
		'''
		columName = [["Etablissement",""],["Filière",""],["Classement",""],
					["Réponse",""],["Numéro",""],["Nom","last_name"],
					["Prénom","first_name"],["Deuxième prénom","middle_name"],
					["Civilité",""],["Sexe","gender","s",["M","Male"],["F","Female"]],
					["Date de naissance","date_of_birth","d","dd/mm/yyyy"],
					["Ville de naissance",""],["Département de naissance",""],
					["Pays de naissance",""],
					["Nationalité",""],["Adresse 1","address_line_1"],
					["Adresse 2","address_line_2"],["Adresse 3",""],
					["Code postal","pincode"],["Commune","city"],
					["Pays",""],["Téléphone",""],["Téléphone portable",""]]

		file_path = os.getcwd()+get_site_path()[1:].encode('utf8') + self.import_file
		wb = load_workbook(filename=file_path, read_only=True)
		ws = wb.active

		start = 2
		for i, row in enumerate(ws.iter_rows(min_row=start)):

			try:
				error = False
#				program_doc = frappe.get_doc("Program", "BTSESF1)
				new_doc = frappe.new_doc("Student Applicant")
				new_doc.program = "BTS Économie Sociale Familiale"
				for j, cell in enumerate(row):
					if j < 21 and columName[j][1] <> "":
#						new_doc.last_name = cell.value
						if len(columName[j]) <= 2:
							setattr(new_doc, columName[j][1], cell.value)
						else:
							if columName[j][2] == "s":
								if columName[j][2][0] == cell.value:
									setattr(new_doc, columName[j][1], columName[j][2][1])
								elif columName[j][3][0] == cell.value:
									setattr(new_doc, columName[j][1], columName[j][3][1])
							elif columName[j][2] == "d":
#								setattr(new_doc, columName[j][1], formatdate(cell.value, columName[j][3]))
								setattr(new_doc, columName[j][1], babel.dates.parse_date(cell.value, locale='fr_FR'))
#								parse_date('01.04.2004', locale='de_DE')
#					if j == 6:
#						new_doc.first_name = cell.value
				new_doc.insert()
				new_doc.save()

			except Exception, e:
				error = True
				if new_doc:
					frappe.errprint(new_doc if isinstance(new_doc, dict) else new_doc.as_dict())
				frappe.errprint(frappe.get_traceback())

			finally:
				frappe.local.message_log = []

			if error:
				frappe.db.rollback()
			else:
				frappe.db.commit()
